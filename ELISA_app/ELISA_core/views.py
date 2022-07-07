from django.shortcuts import render
from .models import Plates
import openpyxl
import seaborn as sns
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import scipy.optimize as optimization
from matplotlib.ticker import ScalarFormatter
import statistics
from operator import itemgetter


def Home(request):
    return render(request, 'Home.html')


def Input_data(request):
    try:
        if request.method == 'POST':
            error = 'correct'
            if request.POST.get('Empty database'):
                Plates.objects.all().delete()
                return render(request, 'Input_data.html', {
                    'check': 'correct_emptied',
                })
            if request.POST.get('file_submit'):
                error = file_data(request)
            if error == "file" or error == "extension":
                return render(request, 'Input_data.html', {
                    'check': error,
                })
            else:
                return render(request, 'Input_data.html', {
                    'check': 'correct',
                })
        else:
            return render(request, 'Input_data.html')
    except:
        return render(request, 'Input_data.html', {
            'check': 'false',
        })


def file_data(request):
    if request.FILES.getlist('my_file') == []:
        return "file"
    for file in request.FILES.getlist('my_file'):
        if str(file).split('.')[1] not in ['txt', 'xlsx', 'xls']:
            return 'extension'
    for file in request.FILES.getlist('my_file'):
        if str(file).split('.')[1] == 'txt':
            data = file.readlines()
            data_string = formatting_txt(data, 2)
            database(data_string, file)
        elif str(file).split('.')[1] == 'xlsx':
            data_string = formatting_xlsx(file)
            database(data_string, file)



def formatting_txt(data, counter):
    lines, data_string, formatted_data = list(), "", list()
    for i in data[:-1]:
        if counter == 1:
            lines.append(i.strip())
        elif counter == 2:
            lines.append(i.strip().decode('utf-8'))
    for j in lines:
        line = j.split('\t')
        formatted_data.append(line)
    formatted_data[1].insert(0, '#')
    for i in formatted_data:
        for j in i:
            data_string += j + "="
    return data_string


def formatting_xlsx(file_name):
    wb = openpyxl.load_workbook(file_name)
    active_sheet = wb.active
    excel_data, data_string = list(), ""
    for row in active_sheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    del excel_data[1][0]
    del excel_data[0][1:]
    excel_data[1].insert(0, '#')
    for i in excel_data:
        for j in i:
            data_string += j + "="
    return data_string


def database(data_string, file):
    split = data_string.split('=')
    plates_instance = Plates.objects.create(
        id=file,
        name=str(split[0]),
        data=data_string
    )


totaal = []
check = ''


def Plate_layout(request):
    global check, totaal
    if request.method == 'POST':
        if request.POST.get('file_submit'):
            totaal = []
            if request.FILES.getlist("my_file") == []:
                check = 'error'
                return render(request, 'Plate_layout.html', {
                    'check': check, 'totaal': totaal,
                })
            excel_data = Plate_layout_1(request)
            totaal = Plate_layout_2(excel_data)
            check = 'go'
            return render(request, 'Plate_layout.html', {
                'totaal': totaal,
                'check': check,
            })
        if request.POST.get('standaard_input'):
            Plate_layout_3(request)
            check = 'go'
            return render(request, 'Plate_layout.html', {
                'totaal': totaal, 'check': check, })
    else:
        return render(request, 'Plate_layout.html', {
            'totaal': totaal, 'check': check, })


def Plate_layout_1(request):
    excel_file = request.FILES["my_file"]
    wb = openpyxl.load_workbook(excel_file)
    active_sheet = wb.active
    excel_data = list()
    for row in active_sheet.iter_rows():
        row_data = list()
        for cell in row:
            if type(cell.value) == float:
                row_data.append(str(round(cell.value)))
            else:
                row_data.append(str(cell.value))
        excel_data.append(row_data)
    return excel_data


def Plate_layout_2(excel_data):
    temp, counter = [], 0
    for i in excel_data:
        i = [e for e in i if e not in ('None')]
        if len(i) != 0:
            if counter == 1:
                i.insert(0, '#')
            temp.append(i)
            counter += 1
            if counter == 10:
                totaal.append(temp)
                counter = 0
                temp = []
    return totaal


def Plate_layout_3(request):
    values = request.POST.get('standaard')
    counter, counter2 = 0, 0
    for i in totaal:
        for j in i[2:]:
            if counter2 != 7:
                j[1] = float(values)
                j[2] = float(values)
                values = float(values) / 2
            elif counter2 == 7:
                j[1] = '#'
                j[2] = '#'
            counter2 += 1
        counter += 1
        values = request.POST.get('standaard')
        counter2 = 0


end_dilution = []

# je kan hier nog een enkele waarde aanpassen door per regel te checken
# wat voor letter is ingevoerd en dan de positie pakken
def Dilutions(request):
    global end_dilution
    if request.method == 'POST':
        if request.POST.get('dilution_submit'):
            dilution = request.POST.get('dilution')
            row_names = ["A", "B", "C", "D", "E", "F", "G", "H"]
            end_list = [["#", "1", "2", "3", "4", "5", "6", "7", "8", "9",
                         "10", "11", "12"]]
            for i in range(8):
                temp = []
                temp = Dilutions_1(i, temp, row_names, dilution)
                end_list.append(temp)
            end_dilution = end_list
            return render(request, 'Dilutions.html', {
                "end_list": end_list
            })
    return render(request, 'Dilutions.html', {
        "end_list": end_dilution})


def Dilutions_1(i, temp, row_names, dilution):
    for x in range(13):
        if i == 0:
            if x == 0:
                temp.append(row_names[i])
            elif x == 1 or x == 2:
                temp.append("1")
            else:
                temp.append(dilution)
        else:
            if x == 0:
                temp.append(row_names[i])
            elif x == 1 or x == 2:
                temp.append("1")
            else:
                temp.append(dilution)
    return temp

dictionary = {}
HD = ''
delete = []
points_dictionary = {}


def Visualize_data(request):
    try:
        global dictionary
        global HD
        global delete
        if request.method == 'POST':
            HD = request.POST['HD']
            bottom = request.POST.getlist('top')
            top = request.POST.getlist('bottom')
            counter = 0
            for keys in dictionary:
                points_dictionary[keys] = [top[counter], bottom[counter]]
                counter += 1
            delete = request.POST.getlist('delete')
        data = Plates.objects.values()
        counter = 0
        nested = []
        temp = []
        for i in data:
            name = i['id']
            lines = i['data'].split('=')[:-1]
            number1 = lines[106].replace(',', '.')
            number2 = lines[107].replace(',', '.')
            calculation = ((float(number1) + float(number2))/2)
            mean = round(calculation, 3)
            for j in lines[1:]:
                if ',' in j:
                    new = float(j.replace(',', '.')) - mean
                    c_color = 3 - new
                    color = round(c_color)*85
                    DCO = round(new, 3)
                    temp.append([DCO, (color, 255, color)])
                elif '.' in j:
                    new = float(j) - mean
                    c_color = 3 - new
                    color = round(c_color)*85
                    DCO = round(new, 3)
                    temp.append([DCO, (color, 255, color)])
                else:
                    temp.append([j, (255, 255, 255)])
                counter += 1
                if counter == 13:
                    nested.append(temp)
                    print(temp)
                    counter = 0
                    temp = []
            dictionary[name] = nested
            nested = []
        if totaal != []:
            create_graph(dictionary)
        return render(request, 'Visualize_data.html', {
            'dictionary': dictionary,
        })
    except:
        return render(request, 'Error.html', {
            'error': 'An error occurred, please be sure to load in the plate layout file and choose a ST value on the '
                     'Plate Layout page.',
        })


mean_ST_dictionary = {}


def create_graph(dictionary):
    conc = totaal[0][2][1]
    x_list = [conc]
    for i in range(6):
        conc = float(conc)/2
        x_list.append(conc)
    y_list = []
    temp = []
    for values in dictionary.values():
        for elements in values[1:-1]:
            mean = ((float(elements[1][0]) + float(elements[2][0]))/2)
            temp.append(round(mean, 3))
        y_list.append(temp)
        temp = []
    global mean_ST_dictionary
    counter = 0
    for keys in dictionary:
        mean_ST_dictionary[keys] = y_list[counter]
        counter += 1
    counter = 0
    for key in dictionary:
        guess = [1, 1, 1, 1, 1]
        params, params_coveriance = optimization.curve_fit(formula, x_list, y_list[counter], guess)
        intermediate_list(key, params)
        x_min, x_max = np.amin(x_list), np.amax(x_list)
        xs = np.linspace(x_min, x_max, 1000)
        plt.scatter(x_list, y_list[counter])
        plt.plot(xs, formula(xs, *params))
        plt.xscale('log')
        plt.grid()
        ax = plt.gca()
        plt.xticks([1.0, 10, 100])
        ax.xaxis.set_major_formatter(ScalarFormatter())
        plt.savefig('ELISA_core/static/images/' + str(key) + '.png')
        plt.close()
        counter += 1

def formula(x, A, B, C, D, E):
    return D + (A-D)/(np.power((1 + np.power((x/C), B)), E))

mean = 0
std = 0
mean2 = 0
std2 = 0
check_cut_off = 'false'
cut_data = []
outlier_value = 0.0
cut_off_value = 0.0

def Cut_off(request):
    """
    Input:

    Output:

    Function:

    """
    try:
        global mean
        global std
        global mean2
        global std2
        global cut_data
        global check_cut_off
        global outlier_value
        global cut_off_value
        cut_dict = {}
        if request.method == 'POST':
            if request.POST.get('outlier_submit'):
                input1 = request.POST.get('input1')
                input2 = request.POST.get('input2')
                outlier_value = (float(input1) * mean) + (float(input2) * std)
                outlier_value = round(outlier_value, 3)
                new_y_list = []
                for data in cut_data:
                    if data < outlier_value:
                        new_y_list.append(data)
                cut_dict['New_OD'] = new_y_list
                mean2 = round(statistics.mean(new_y_list), 3)
                std2 = round(statistics.stdev(new_y_list), 3)
                df = pd.DataFrame(data=cut_dict)
                ax = sns.swarmplot(data=df, y="New_OD")
                ax = sns.boxplot(data=df, y="New_OD", color='white')
                plt.savefig('ELISA_core/static/images/' + 'swarmplot2.png')
                plt.close()
                check_cut_off = 'true'
                return render(request, 'Cut_off.html', {
                    'mean': mean,
                    'std': std,
                    'mean2': mean2,
                    'std2': std2,
                    'check': check_cut_off,
                    'outlier_value': outlier_value,
                    'cut_off_value': cut_off_value,
                })
            elif request.POST.get('cut_off_submit'):
                input3 = request.POST.get('input3')
                input4 = request.POST.get('input4')
                cut_off_value = (float(input3) * mean2) + (float(input4) * std2)
                cut_off_value = round(cut_off_value, 3)
                return render(request, 'Cut_off.html', {
                    'mean': mean,
                    'std': std,
                    'mean2': mean2,
                    'std2': std2,
                    'check': check_cut_off,
                    'outlier_value': outlier_value,
                    'cut_off_value': cut_off_value,
                })
        elif cut_data == []:
            for i in dictionary[HD][1:]:
                for g in i[3:8]:
                    cut_data.append(g[0])
            cut_data.pop(0)
            cut_dict["OD"] = cut_data
            mean = round(statistics.mean(cut_data), 3)
            std = round(statistics.stdev(cut_data), 3)
            df = pd.DataFrame(data=cut_dict)
            ax = sns.swarmplot(data=df, y="OD")
            ax = sns.boxplot(data=df, y="OD", color='white')
            plt.savefig('ELISA_core/static/images/' + 'swarmplot.png')
            plt.close()
            return render(request, 'Cut_off.html', {
                'mean': mean,
                'std': std,
                'check': check_cut_off,
                'outlier_value': outlier_value,
                'cut_off_value': cut_off_value,
            })
        return render(request, 'Cut_off.html', {
            'mean': mean,
            'std': std,
            'mean2': mean2,
            'std2': std2,
            'check': check_cut_off,
            'outlier_value': outlier_value,
            'cut_off_value': cut_off_value,
        })
    except:
        return render(request, 'Error.html', {
            'error': 'An error occurred, please be sure to select the plate with the healthy donor data on the '
                     'Visualize data page.'
        })


def formula2(y, A, B, C, D, E):
    """
    Input:

    Output:

    Function:

    """
    return C*(np.power((np.power(((A-D)/(-D+float(y))), (1/E))-1), (1/B)))


end_result = {}
lower = 0.0
upper = 0.0


def Intermediate_result(request):
    """
    Input:
        - request: Catches submits form template.
        - end_result: An empty dictionary.
        - lower: The number zero.
        - upper: The number zero.
       - points_dictionary: A dictionary with as key the name of the plate and as value a list of the chosen lower and
                            upper points of that plate.
        - intermediate_dictionary: A dictionary with as keys the plates names and the value a nested list. With in it
          a sample ID and the calculated au/ml.
        - delete: A list with selected plate names which are not allowed in end results.
        - mean_ST_dictionary: A dictionary with as key the names of the plates and the value the average score of ODs.
        - end_dilution: A nested list with the dilution of the plates. the values are all numbers with as type string.
    Output:
        - end_result: A dictionary with as key the name of the plates and the values a nested list. the nested list
                      have as values first a sample id, second the au/ml, and third a 1 or 2 or 3.
        - lower: An au/ml score from the lowest chosen value.
        - upper: An au/ml score from the highest chosen value.
    Function:
        - The fuction first checks which plates it should not to be looking if the plate names in intermediate_dictionary
          are in delete. Then it looks what the values are of the highest chosen value and lowest chosen value. After that
          it will look if the values from end_result are smaller or bigger then lower or upper. If the value is smaller
          than lower it gets a 1, if higher than upper it gets a 3. If nether than it gets a 2. When the list is filled
          it gets sorted by sample ID and everything gets put into one list.
    """
    try:
        global end_result
        global lower
        global upper
        end_result = {}
        for key, values in intermediate_dictionary.items():
            if key not in delete:
                end_result[key] = values
        temp0 = []
        temp1 = []
        temp2 = []
        temp3 = []
        temp4 = []
        for key, values in end_result.items():
            mean_ST_dictionary[key].reverse()
            top = mean_ST_dictionary[key][int(points_dictionary[key][1]) - 1]
            bot = mean_ST_dictionary[key][int(points_dictionary[key][0]) - 1]
            string_top = formula2(top, *params_dictionary[key]) * int(end_dilution[3][3])
            string_bot = formula2(bot, *params_dictionary[key]) * int(end_dilution[3][3])
            for value in values:
                if type(value[1]) == str:
                    if len(value) == 2:
                        if float(value[1]) < bot:
                            temp0.append([value[0]] + ['<' + str(round(string_bot, 3))] + [1])
                        else:
                            temp4.append([value[0]] + ['>' + str(round(string_top, 3))] + [3])
                    else:
                        value[2] = 1
                        temp0.append(value)
                elif int(value[1]) <= float(string_bot):
                    if len(value) == 2:
                        temp1.append(value + [1])
                    else:
                        value[2] = 1
                        temp1.append(value)
                elif int(value[1]) >= float(string_top):
                    if len(value) == 2:
                        temp3.append(value + [3])
                    else:
                        value[2] = 3
                        temp3.append(value)
                else:
                    if len(value) == 2:
                        temp2.append(value + [2])
                    else:
                        value[2] = 2
                        temp2.append(value)
            mean_ST_dictionary[key].reverse()
        sorted_temp1 = sorted(temp1, key=itemgetter(1))
        sorted_temp2 = sorted(temp2, key=itemgetter(1))
        sorted_temp3 = sorted(temp3, key=itemgetter(1))
        lower = sorted_temp2[0][1]
        upper = sorted_temp2[-1][1]
        complete_list = temp0 + sorted_temp1 + sorted_temp2 + sorted_temp3 + temp4
        if request.method == 'POST':
            if request.POST.get('limit_submit'):
                lower = request.POST.get('lower')
                upper = request.POST.get('upper')
        return render(request, 'Intermediate_result.html', {
            'complete_list': complete_list,
        })
    except:
        return render(request, 'Error.html', {
            'error': 'An error occurred, please make sure you have selected the the healthy donor plate and confirming '
                     'your preferences on the visualize Data page.'
        })


intermediate_dictionary = {}
params_dictionary = {}


def intermediate_list(key, params):
    """
    Input:
        - key: A string with the name of the plate.
        - params: The optimal values for the curve fit. A: minimal asymptote, B: steepness, C: inflection point
                 D: maximal asymptote, E: Asymmetry factor.
        - intermediate_dictionary: An empty dictionary.
        - params_dictionary: An Empty dictionary.
        - totaal: Nested list with submitted data from a plate_layout file.
        - end_dilution: A nested list with the dilution of the plates. the values are all numbers with as type string.
    Output:
        - intermediate_dictionary: A dictionary with as keys the plates names and the value a nested list. With in it
          a sample ID and the calculated au/ml.
        - params_dictionary: In this dictionary the keys are the name of the plates and the values the params.
    Function:
            - First the function takes a look which plate from visualize data belongs with the plate from
              plate-layout by comparing the number in de plate name. then it save the params in the params_dictionary
              and calls the formula2 function, so the result can be multiplied by the dilution score to get the au/ml.
    """
    global intermediate_dictionary
    for options in range(len(totaal)):
        num1 = int(''.join(filter(str.isdigit, key)))
        num2 = int(''.join(filter(str.isdigit, totaal[options][0][0])))
        if num1 == num2:
            position = options
    dilution = end_dilution[3][3]
    list1 = []
    for i, j in dictionary.items():
        if i == key:
            params_dictionary[key] = params
            for values in range(len(j)):
                if values != 0:
                    for value in range(len(j[values])):
                        if value != 0 and value != 1 and value != 2:
                            result = formula2(j[values][value][0], *params)
                            if np.isnan(result):
                                result = str(j[values][value][0])
                            else:
                                result *= int(dilution)
                                result = round(result, 3)
                            list1.append([totaal[position][values + 1][value], result])
            intermediate_dictionary[i] = list1


final_dictionary = {}
final_list = []
cut_off_value_au = 0


def End_results(request):
    """
    Input:
        - request: Catches submits form template.
        - final_dictionary: An empty dictionarty list.
        - final_list: An empty list.
        - cut_off_value_au: The number zero.
        - dictionary: A dictionary with as key the plate names and the value a nested list with in it the OD and RBG code.
        - cut_off_value: It has contains a float which is the OD given by the cut_off function
        - params_dictionary: In this dictionary the keys are the name of the plates and the values the params.
        - end_dilution: A nested list with the dilution of the plates. the values are all numbers with as type string.
        - end_result: A dictionary with as key the name of the plates and the values a nested list. the nested list
                      have as values first a sample id, second the au/ml, and third a 1 or 2 or 3.
        - lower: An au/ml score.
        - delete: A list with selected plate names which are not allowed in end results.
    Output:
        - final_dictionary: The dictionary is now filled and has as key sampleID which start with 1 and goes up by 1
                            with every now result. the values are a list with as first a sample id, second an 1 or 2,
                            and third the au/ml.
        - final_list: A list with as first a sample id, second an 1 or 2 and third the au/ml.
        - cut_off_value_au: An au/ml, calculated from the OD from cut_off_value and the params from params_dictionary.
        - end_result: A dictionary with as key the name of the plates and the values a nested list. the nested list
                      have as values first a sample id, second the au/ml, and third a 1 or 2 or 3.
    Function:
        - The function first checks if any button was pressed, if their were any buttons pressed it then check which.
          After checking which buttons where pressed it will fill up the final_dictionary and final_list by checking
          if they pass any off the requirement given by the if-statement. If all the requirements are met then the list
          is given an 1, if they are not met then the list gets an 0. After the list is filled and sorted
          the list is given to the render.
    """
    try:
        global final_list
        global cut_off_value_au
        global final_dictionary
        if request.method == 'POST':
            if request.POST.get('Empty database'):
                Plates.objects.all().delete()
            if request.POST.get('download'):
                file_name = request.POST.get('File_name')
                textfile = open("../Download_files/" + file_name + ".txt", "w")
                for elements in final_list:
                    for element in elements:
                        textfile.write(str(element) + "\t")
                    textfile.write("\n")
                textfile.close()
            if request.POST.get('update_table'):
                final_dictionary = {}
                OD_multiplier = request.POST.get('OD_multiplier')
                if len(end_result[HD][0]) == 2:
                    for keys, values in dictionary.items():
                        if keys == HD:
                            params = params_dictionary[HD]
                            cut_off_value_au = formula2(float(cut_off_value), *params) * int(end_dilution[3][3])
                        if keys not in delete:
                            counter = 0
                            for OD_list in values[1:]:
                                for OD in OD_list[3:]:
                                    end_result[keys][counter].append(OD[0])
                                    counter += 1
                sampleID = 1
                final_list = []
                for keys, values in end_result.items():
                    if keys != HD:
                        counter = 0
                        counter2 = 0
                        for elements in values:
                            if counter < 5:
                                if float(elements[1]) >= float(lower):
                                    if elements[1] >= float(cut_off_value_au):
                                        if (values[counter2][2])/(values[counter2 + 5][2]) >= int(OD_multiplier):
                                            final_dictionary[sampleID] = [elements[0], 1, round(elements[1])]
                                if sampleID not in final_dictionary:
                                    if float(elements[1]) < float(lower):
                                        final_dictionary[sampleID] = [elements[0], 0, '<' + str(lower)]
                                    else:
                                        final_dictionary[sampleID] = [elements[0], 0, round(float(elements[1]))]
                                sampleID += 1
                            counter += 1
                            counter2 += 1
                            if counter == 10:
                                counter = 0
                for i, lists in final_dictionary.items():
                    final_list.append(lists)
                final_list = sorted(final_list, key=itemgetter(0))
        return render(request, 'End_results.html', {
            'final_list': final_list,
            'lower': lower,
            'cut_off_value': round(cut_off_value_au),
        })
    except:
        return render(request, 'Error.html', {
            'error': 'An error occurred, please make sure you have submitted all the settings on previous pages.'
        })
